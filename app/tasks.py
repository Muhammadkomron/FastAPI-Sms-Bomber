import io
from datetime import datetime

from openpyxl.reader.excel import load_workbook

from app import crud, schemas
from app.choices import LogStatus
from app.models import Notification
from app.services import EskizClient
from app.config import logger


async def process_excel_data(file_content: bytes):
    workbook = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    sheet = workbook.active
    count = 0
    for row in sheet.iter_rows(values_only=True, min_row=2):
        count += 1
        date_str = str(int(row[4])) if type(row[4]) in (float, int) else row[4]
        date_str = date_str.replace('(', '').replace(')', '').replace('(', '').replace(')', '').replace('*', '') \
            .replace(',', '.').replace(' ', '').replace('.yil', '')
        date_str = date_str[:-2] if date_str.endswith('.') else date_str
        date_str = date_str[1:] if date_str.startswith('.') else date_str
        birth_date = None
        try:
            birth_date = datetime.strptime(date_str, '%d.%m.%Y').date()
        except ValueError:
            try:
                birth_date = datetime.strptime(date_str, '%m.%d.%Y').date()
            except ValueError:
                pass
        phone_number = str(int(row[6])) if type(row[6]) in (float, int) else row[6]
        phone_number = phone_number.replace(' ', '').replace('-', '').replace('.', '').replace('(', '') \
            .replace(')', '').replace('*', '').replace(',', '').replace('yoʻq', '')
        phone_numbers = phone_number.split('+')
        phone_numbers = list(filter(None, phone_numbers))
        if '+' in phone_number and len(phone_numbers) > 1:
            for i in phone_numbers:
                if 9 > len(i) > 12:
                    continue
                if count == 1:
                    logger.info(f"count: {i}, phone_number.split('+'): {phone_number.split('+')}")
                student = schemas.StudentCreateUpdate(
                    phone_number=f'998{i}' if len(i) == 9 else i,
                    username=str(row[5]),
                    full_name=str(row[1]),
                    birth_date=birth_date,
                )
                await crud.get_or_create_student(student)
            continue
        phone_number = phone_number.replace('+', '')
        if 9 > len(phone_number) > 12:
            continue
        if len(phone_number) == 9:
            phone_number = f'998{phone_number}'
        student = schemas.StudentCreateUpdate(
            phone_number=f'998{phone_number}' if len(phone_number) == 9 else phone_number,
            username=str(row[5]),
            full_name=str(row[1]),
            birth_date=birth_date,
        )
        await crud.get_or_create_student(student)
    logger.info(f'Background task successfully finished. Inserted {count} Students!')


async def process_notification(model: Notification):
    offset = 0
    limit = 10
    while True:
        students = await crud.get_students(offset=offset, limit=limit)
        for student in students:
            status = LogStatus.PENDING
            try:
                EskizClient().send_sms(student.phone_number, model.message)
                logger.info(f'id: {student.id} phone_number: {student.phone_number}')
            except Exception as e:
                status = LogStatus.FAILED
                logger.info(f'Sms sending error: {student.phone_number} {str(e)}')
            finally:
                log = schemas.LogBase(student_id=student.id, notification_id=model.id, status=status)
                await crud.create_log(log)
        if len(students) != limit:
            break
        offset += 10
    logger.info('Background task successfully finished!')
